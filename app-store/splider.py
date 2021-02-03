#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

import requests
# import urllib.request
# import re
# import xlsxwriter
import openpyxl
import json
import inquirer
# import os
import time
import os

apps = {}

def search_app_id(app): 
  url = "http://itunes.apple.com/search?term=" + app + "&entity=software"
  r = requests.get(url)
  html = r.content
  html_doc = str(html, 'utf-8')
  data = json.loads(html_doc)
  resultCount = data['resultCount']
  results = data['results']
  print(app + ' Find ' + str(resultCount) + 'result(s)')

  for i in range(resultCount):
    app_name = results[i]['trackName']
    app_id = results[i]['trackId']
    apps[app_id] = app_name
    print('name: ' + app_name, 'id: ' + str(app_id))

def save_data_to_ws(wb, ws, data, app):
  app_id = app['app_id']
  app_name = app['app_name']

  for i in data:
    name = i['author']['name']['label']
    rate = i['im:rating']['label']
    user_id = i['id']['label']
    content = i['content']['label']
    version = i['im:version']['label']
    
    # print(name, rate, user_id,  content)
    write_ws(ws, [
      app_id,
      app_name,
      name,
      rate,
      user_id, 
      content,
      version
    ], False)  
  wb.save('app_store.xlsx')

def fetch_app_comment(app_id):
  comments = []

  for j in range(1, 11):
    url = 'https://itunes.apple.com/rss/customerreviews/page=' + str(j) + '/id=' + str(app_id) + '/sortby=mostrecent/json?l=en&&cc=cn'
    print('当前地址：' + url)
    r = requests.get(url)
    
    if r.status_code == 200:
      html = r.content
      html_doc = str(html, 'utf-8')
      entry = json.loads(html_doc)['feed'].get('entry') or []
      data =  entry if isinstance(entry, list) else []
      comments = comments + data
      print('一共有: ' + str(len(comments)) + '条数据')

    time.sleep(2)
  
  return comments
def get_app_ids():
  is_continue = True
  ids = []

  while is_continue:
    app_id = input("请输入app id: \n")
    
    if not int(app_id) in apps:
      break
    

    app_name = apps[int(app_id)]
    print('应用名称: ' + app_name)

    ids.append({ 'app_id': app_id, 'app_name': app_name })

    is_continue = is_ok('还要继续抓取其他app的评论数据吗')
  return ids

def start_fetch(wb, ws):
  ids = get_app_ids()

  for i in range(len(ids)):
    data = fetch_app_comment(ids[i]['app_id'])
    save_data_to_ws(wb, ws, data , ids[i])

def is_ok(msg):
  questions = [
    inquirer.List('is_ok',
                  message=msg,
                  choices=['yes', 'no'],
    ),
  ]
  answers = inquirer.prompt(questions)

  return True if answers['is_ok'] == 'yes' else False

def input_name(msg, default):
  questions = [
    inquirer.Text('name', message=msg)
  ]
  answers = inquirer.prompt(questions)
  name = answers['name'] or default

  return name

def select(msg, choices):
  questions = [
      inquirer.List('choices',
                    message=msg,
                    choices=choices
      )
  ]
  answers = inquirer.prompt(questions)
  choice = answers['choices']

  return choice

def write_ws(ws, fields, row):
  max_row = ws.max_row

  for i in range(len(fields)):
    ws.cell(row= row if row else max_row + 1 , column=i+1, value=fields[i])  

def main():

    name = input("请输入应用名称:")
    search_app_id(name)

    comment_fields = [
      'APP ID',
      'APP 名称',
      '昵称',
      '评分',
      '用户id',
      '评论',
      '版本'
    ]
    
    if os.path.exists('app_store.xlsx'):
      wb = openpyxl.load_workbook('app_store.xlsx')
      

      if is_ok('是否创建新表'):
        sheet_name = input_name('请输入新表名称?','comment')
        ws = wb.create_sheet(sheet_name)
        write_ws(ws, comment_fields, 1)
      
      else:
        sheets = wb.sheetnames
        ws = wb[select('选择表', sheets)]

      start_fetch(wb, ws)
    else:
      # Workbook init
      wb = openpyxl.Workbook()
      ws = wb.active

      ws.title = input_name('请输入新表名称?', 'comment')
      write_ws(ws, comment_fields, 1)

      start_fetch(wb, ws)

    
    print('Done!')

if __name__ == '__main__':
    main()
